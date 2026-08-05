"""Import adapter for raw ERA / ERP / CRM export files.

Boundary:
    This service only parses files, reads sheets, normalizes cell values, and returns
    raw fact rows. It must not create risk judgments, task leads, advice, or report
    explanations. All trend, cross-validation, and task generation logic belongs to
    the downstream data import services.

V12.2 update:
    Preserve row coordinates and sheet matrices so the report layout agent can split
    one sheet into multiple blocks: 商品经营区、店铺汇总区、流量来源区、暂存区。
"""

from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Tuple

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".csv", ".json"}
ADAPTER_VERSION = "12.2.2"


def _extension(filename: str | None) -> str:
    suffix = Path(filename or "upload").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError("仅支持 Excel / CSV / JSON 文件：.xlsx、.xlsm、.xls、.csv、.json")
    return suffix


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _is_empty(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        return value.strip()
    return value


def _dedupe_headers(headers: List[Any]) -> List[str]:
    seen: Dict[str, int] = {}
    result: List[str] = []
    for index, header in enumerate(headers):
        text = str(_normalize_cell(header) or "").strip()
        if not text:
            text = f"未命名字段{index + 1}"
        count = seen.get(text, 0) + 1
        seen[text] = count
        result.append(text if count == 1 else f"{text}_{count}")
    return result


def _matrix_payload(matrix: List[List[Any]]) -> List[Dict[str, Any]]:
    payload: List[Dict[str, Any]] = []
    for row_index, row in enumerate(matrix, start=1):
        payload.append({
            "rowIndex": row_index,
            "cells": [_normalize_cell(cell) for cell in row],
            "isEmpty": not any(not _is_empty(cell) for cell in row),
        })
    return payload


def _matrix_to_rows(matrix: List[List[Any]], *, source_sheet: str | None = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    non_empty = [(index, row) for index, row in enumerate(matrix, start=1) if any(not _is_empty(cell) for cell in row)]
    if not non_empty:
        return [], []
    header_row_index, header_row = non_empty[0]
    headers = _dedupe_headers(header_row)
    column_map = {header: index + 1 for index, header in enumerate(headers)}
    rows: List[Dict[str, Any]] = []
    for source_row_index, raw in non_empty[1:]:
        item: Dict[str, Any] = {}
        for index, header in enumerate(headers):
            item[header] = _normalize_cell(raw[index]) if index < len(raw) else ""
        if source_sheet:
            item["__source_sheet"] = source_sheet
        item["__source_row_index"] = source_row_index
        item["__source_header_row_index"] = header_row_index
        item["__source_column_map"] = column_map
        if any(not _is_empty(value) for key, value in item.items() if not str(key).startswith("__")):
            rows.append(item)
    return headers, rows


def _parse_csv(content: bytes) -> Dict[str, Any]:
    text = _decode_text(content)
    reader = csv.reader(io.StringIO(text))
    matrix = [list(row) for row in reader]
    headers, rows = _matrix_to_rows(matrix, source_sheet="CSV")
    return {
        "format": "csv",
        "rows": rows,
        "sheetRows": {"CSV": rows},
        "sheetMatrices": {"CSV": _matrix_payload(matrix)},
        "totalRows": len(rows),
        "sheetCount": 1,
        "sheets": [{"sheetName": "CSV", "headers": headers, "rowCount": len(rows), "rawRowCount": len(matrix)}],
    }


def _parse_json(content: bytes) -> Dict[str, Any]:
    payload = json.loads(_decode_text(content))
    if isinstance(payload, list):
        rows = [item for item in payload if isinstance(item, dict)]
    elif isinstance(payload, dict):
        source_rows = payload.get("rows") or payload.get("data") or []
        rows = [item for item in source_rows if isinstance(item, dict)] if isinstance(source_rows, list) else []
    else:
        rows = []
    if not rows:
        raise ValueError("JSON 需要是对象数组，或包含 rows/data 数组。")
    headers: List[str] = []
    for row in rows[:20]:
        for key in row.keys():
            if key not in headers:
                headers.append(str(key))
    column_map = {header: index + 1 for index, header in enumerate(headers)}
    for index, row in enumerate(rows, start=2):
        row.setdefault("__source_sheet", "JSON")
        row.setdefault("__source_row_index", index)
        row.setdefault("__source_header_row_index", 1)
        row.setdefault("__source_column_map", column_map)
    return {
        "format": "json",
        "rows": rows,
        "sheetRows": {"JSON": rows},
        "sheetMatrices": {"JSON": [{"rowIndex": 1, "cells": headers, "isEmpty": False}]},
        "totalRows": len(rows),
        "sheetCount": 1,
        "sheets": [{"sheetName": "JSON", "headers": headers, "rowCount": len(rows), "rawRowCount": len(rows) + 1}],
    }


def _parse_xlsx(content: bytes) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("缺少 openpyxl 依赖，无法解析 .xlsx / .xlsm 文件。") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows: List[Dict[str, Any]] = []
    sheet_rows: Dict[str, List[Dict[str, Any]]] = {}
    sheet_matrices: Dict[str, List[Dict[str, Any]]] = {}
    sheets: List[Dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
        headers, current_rows = _matrix_to_rows(matrix, source_sheet=worksheet.title)
        if not headers and not current_rows:
            continue
        rows.extend(current_rows)
        sheet_rows[worksheet.title] = current_rows
        sheet_matrices[worksheet.title] = _matrix_payload(matrix)
        sheets.append({"sheetName": worksheet.title, "headers": headers, "rowCount": len(current_rows), "rawRowCount": len(matrix)})
    if not rows:
        raise ValueError("Excel 文件没有读取到有效数据行。")
    return {"format": "xlsx", "rows": rows, "sheetRows": sheet_rows, "sheetMatrices": sheet_matrices, "totalRows": len(rows), "sheetCount": len(sheets), "sheets": sheets}


def _parse_xls(content: bytes) -> Dict[str, Any]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("缺少 xlrd 依赖，无法解析 .xls 文件。请安装依赖或另存为 .xlsx。") from exc

    book = xlrd.open_workbook(file_contents=content)
    rows: List[Dict[str, Any]] = []
    sheet_rows: Dict[str, List[Dict[str, Any]]] = {}
    sheet_matrices: Dict[str, List[Dict[str, Any]]] = {}
    sheets: List[Dict[str, Any]] = []
    for sheet in book.sheets():
        matrix: List[List[Any]] = []
        for row_index in range(sheet.nrows):
            row: List[Any] = []
            for col_index in range(sheet.ncols):
                cell = sheet.cell(row_index, col_index)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, book.datemode)
                    except Exception:
                        pass
                row.append(value)
            matrix.append(row)
        headers, current_rows = _matrix_to_rows(matrix, source_sheet=sheet.name)
        if not headers and not current_rows:
            continue
        rows.extend(current_rows)
        sheet_rows[sheet.name] = current_rows
        sheet_matrices[sheet.name] = _matrix_payload(matrix)
        sheets.append({"sheetName": sheet.name, "headers": headers, "rowCount": len(current_rows), "rawRowCount": len(matrix)})
    if not rows:
        raise ValueError("Excel 文件没有读取到有效数据行。")
    return {"format": "xls", "rows": rows, "sheetRows": sheet_rows, "sheetMatrices": sheet_matrices, "totalRows": len(rows), "sheetCount": len(sheets), "sheets": sheets}


def parse_upload_file(filename: str | None, content: bytes, content_type: str | None = None) -> Dict[str, Any]:
    ext = _extension(filename)
    if not content:
        raise ValueError("上传文件为空。")
    if ext in {".xlsx", ".xlsm"}:
        parsed = _parse_xlsx(content)
    elif ext == ".xls":
        parsed = _parse_xls(content)
    elif ext == ".csv":
        parsed = _parse_csv(content)
    elif ext == ".json":
        parsed = _parse_json(content)
    else:
        raise ValueError("不支持的文件格式。")
    parsed["fileName"] = filename or "upload"
    parsed["contentType"] = content_type or ""
    parsed["adapterVersion"] = ADAPTER_VERSION
    parsed["boundary"] = "raw_fact_rows_only_no_task_judgement_v12_2_row_coordinates_preserved"
    return parsed


def compact_upload_meta(parsed: Dict[str, Any]) -> Dict[str, Any]:
    try:
        from src.services.report_profile_agent_service import build_report_profile

        report_profile = build_report_profile(parsed)
    except Exception as exc:  # profile failure must not block raw parsing
        report_profile = {"version": "12.2.2", "status": "profile_failed", "error": str(exc)}
    return {
        "adapterVersion": parsed.get("adapterVersion"),
        "fileName": parsed.get("fileName"),
        "contentType": parsed.get("contentType"),
        "format": parsed.get("format"),
        "totalRows": parsed.get("totalRows", 0),
        "sheetCount": parsed.get("sheetCount", 0),
        "sheets": parsed.get("sheets", []),
        "reportProfile": report_profile,
        "boundary": parsed.get("boundary"),
    }
