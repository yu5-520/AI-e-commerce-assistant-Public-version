#!/usr/bin/env python3
"""Attest REC-001: ERA 10x3 operating units plus canonical multi-DV history."""
from __future__ import annotations

import argparse
import hashlib
import io
import json
import sys
from pathlib import Path
from typing import Any, Sequence

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402
from src.services import canonical_product_trend_v2_service as trend_bridge  # noqa: E402
from src.services.competition_era_sample_payload_v2 import sample_contract  # noqa: E402
from src.services.competition_sample_report_service import (  # noqa: E402
    SAMPLE_REPORTS,
    SAMPLE_SHEET_ORDER,
    build_competition_sample_xlsx,
)

SCHEMA = "competition.era_recovery_contract.v1"
EXPECTED_ROWS = {"商品经营明细": 30, "店铺经营汇总": 3, "流量来源明细": 150}


def _dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _hash(value: Any) -> str:
    raw = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str
    ).encode("utf-8")
    return "sha256:" + hashlib.sha256(raw).hexdigest()


def _write(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
        encoding="utf-8",
    )


def _period(period: int) -> dict[str, Any]:
    first = build_competition_sample_xlsx(period)
    second = build_competition_sample_xlsx(period)
    workbook = load_workbook(io.BytesIO(first), read_only=True, data_only=True)
    rows = list(workbook["商品经营明细"].iter_rows(values_only=True))
    header = list(rows[0])
    s, p, k = header.index("店铺ID"), header.index("商品ID"), header.index("SKU ID")
    units = {(row[s], row[p], row[k]) for row in rows[1:]}
    products = {row[p] for row in rows[1:]}
    stores = {row[s] for row in rows[1:]}
    sheet_rows = {
        name: max(0, int(workbook[name].max_row or 0) - 1) for name in SAMPLE_SHEET_ORDER
    }
    contract = sample_contract(period)
    signals = SAMPLE_REPORTS.get(period) or []
    assertions = {
        "byteDeterministic": first == second,
        "xlsxMagic": first.startswith(b"PK"),
        "sheetOrder": list(workbook.sheetnames) == list(SAMPLE_SHEET_ORDER),
        "sheetRows": sheet_rows == EXPECTED_ROWS,
        "operatingUnits30": len(units) == 30,
        "globalProducts10": len(products) == 10,
        "stores3": len(stores) == 3,
        "contractOperatingUnits30": contract.get("operatingProductUnitCount") == 30,
        "contractGlobalProducts10": contract.get("globalProductCount") == 10,
        "contractStores3": contract.get("storeCount") == 3,
        "rows183": sum(sheet_rows.values()) == 183,
        "signalFixtureIsThreeOnly": len(signals) == 3,
        "signalFixtureCannotOwnInventory": 30 > len(signals),
    }
    return {
        "period": period,
        "verified": all(assertions.values()),
        "xlsxSha256": "sha256:" + hashlib.sha256(first).hexdigest(),
        "operatingUnitCount": len(units),
        "globalProductCount": len(products),
        "storeCount": len(stores),
        "signalFixtureCount": len(signals),
        "sheetRows": sheet_rows,
        "assertions": assertions,
    }


def _product(
    object_id: str,
    store: str,
    product: str,
    sku: str,
    date: str,
    payment: float,
    roi: float,
    conversion: float,
) -> dict[str, Any]:
    return {
        "objectId": object_id,
        "storeId": store,
        "productId": product,
        "skuId": sku,
        "profileSnapshot": {
            "objectId": object_id,
            "storeId": store,
            "productId": product,
            "skuId": sku,
            "title": "通勤防泼水背包",
            "platform": "天猫",
        },
        "metricSnapshot": {
            "metricDate": date,
            "paymentAmount": payment,
            "roi": roi,
            "conversionRate": conversion,
            "refundRate": 0.03,
            "sourceDataVersions": [],
        },
    }


def _history() -> dict[str, Any]:
    object_id = "product::tianmao::TB-SH-001::P10004::SKU10004-A"
    sibling_id = "product::jingdong::JD-SH-002::P10004::SKU10004-B"
    versions = [
        ("DV-3", "2026-07-02", 4116.42, 3.35, 0.0427),
        ("DV-2", "2026-06-28", 3929.31, 3.35, 0.0426),
        ("DV-1", "2026-06-25", 3555.09, 3.35, 0.0438),
    ]
    snapshots: dict[str, dict[str, Any]] = {}
    for dv, date, payment, roi, conversion in versions:
        snapshots[dv] = {
            "snapshotId": f"SNAP-{dv}",
            "dataVersion": dv,
            "createdAt": f"{date}T23:10:00",
            "updatedAt": f"{date}T23:10:00",
            "products": [
                _product(object_id, "TB-SH-001", "P10004", "SKU10004-A", date, payment, roi, conversion),
                _product(sibling_id, "JD-SH-002", "P10004", "SKU10004-B", date, payment * 2, roi + 1, conversion + 0.01),
            ],
        }

    metadata = [
        {
            "snapshot_id": f"SNAP-{dv}",
            "data_version": dv,
            "set_snapshot_hash": f"sha256:{hashlib.sha256(dv.encode('utf-8')).hexdigest()}",
            "created_at": snapshots[dv]["createdAt"],
            "updated_at": snapshots[dv]["updatedAt"],
        }
        for dv, *_ in versions
    ]

    def fake_slim_snapshot(meta: dict[str, Any], product_id: str, store_id: str | None):
        snapshot = snapshots.get(str(meta.get("data_version") or ""))
        if not snapshot:
            return None
        matched = next(
            (
                item
                for item in snapshot.get("products") or []
                if isinstance(item, dict)
                and trend_bridge._matches(item, product_id, store_id)
            ),
            None,
        )
        if matched is None:
            return None
        return {
            "snapshotId": meta.get("snapshot_id"),
            "dataVersion": meta.get("data_version"),
            "setSnapshotHash": meta.get("set_snapshot_hash"),
            "createdAt": meta.get("created_at"),
            "updatedAt": meta.get("updated_at"),
            "products": [trend_bridge._slim_product(matched)],
        }

    original_metadata = trend_bridge._history_metadata
    original_slim = trend_bridge._slim_snapshot_for_product
    try:
        trend_bridge._history_metadata = lambda limit=120: list(metadata)
        trend_bridge._slim_snapshot_for_product = fake_slim_snapshot
        trend_bridge._CACHE.clear()
        trend = trend_bridge.read_canonical_product_trend(
            object_id, store_id="TB-SH-001", user_id="competition-recovery-probe"
        )

        trend_bridge._slim_snapshot_for_product = lambda meta, product_id, store_id: None
        trend_bridge._CACHE.clear()
        missing = trend_bridge.read_canonical_product_trend(
            "missing", store_id="TB-SH-001", user_id="competition-recovery-probe"
        )
    finally:
        trend_bridge._history_metadata = original_metadata
        trend_bridge._slim_snapshot_for_product = original_slim
        trend_bridge._CACHE.clear()

    recent = trend.get("recentSnapshots") if isinstance(trend.get("recentSnapshots"), list) else []
    assertions = {
        "ready": trend.get("ready") is True,
        "canonicalAuthority": trend.get("snapshotAuthority") == "canonical_product_snapshot_sets_v1",
        "legacyFallbackDisabled": trend.get("legacySnapshotFallbackUsed") is False,
        "boundedSingleProductScan": trend.get("historyScanMode") == "metadata_then_single_row_single_product"
        and trend.get("wholeSnapshotRetention") is False,
        "threeValidSnapshots": _dict(trend.get("observationSummary")).get("validSnapshotCount") == 3,
        "threeDataVersions": [item.get("dataVersion") for item in recent] == ["DV-1", "DV-2", "DV-3"],
        "sameStoreOnly": _dict(trend.get("product")).get("storeId") == "TB-SH-001",
        "sameSkuOnly": _dict(trend.get("product")).get("skuId") == "SKU10004-A",
        "noFabrication": missing.get("recentSnapshots") == []
        and _dict(missing.get("observationSummary")).get("validSnapshotCount") == 0,
    }
    return {
        "verified": all(assertions.values()),
        "objectId": object_id,
        "dataVersions": [item.get("dataVersion") for item in recent],
        "businessDates": [item.get("businessDate") for item in recent],
        "storeId": _dict(trend.get("product")).get("storeId"),
        "skuId": _dict(trend.get("product")).get("skuId"),
        "historyScanMode": trend.get("historyScanMode"),
        "wholeSnapshotRetention": trend.get("wholeSnapshotRetention"),
        "assertions": assertions,
    }


def build_report() -> dict[str, Any]:
    periods = [_period(period) for period in (1, 2, 3)]
    history = _history()
    assertions = {
        "allPeriodsVerified": all(item["verified"] for item in periods),
        "shapeStableAcrossPeriods": {
            (item["operatingUnitCount"], item["globalProductCount"], item["storeCount"])
            for item in periods
        }
        == {(30, 10, 3)},
        "canonicalHistoryVerified": history.get("verified") is True,
    }
    material = {
        "schema": SCHEMA,
        "periods": periods,
        "canonicalHistory": history,
        "assertions": assertions,
    }
    return {**material, "verified": all(assertions.values()), "eraRecoveryHash": _hash(material)}


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        default="dist/competition-three-report-e2e/era-recovery-contract.json",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    report = build_report()
    _write(Path(args.output), report)
    print(
        json.dumps(
            {
                "verified": report["verified"],
                "eraRecoveryHash": report["eraRecoveryHash"],
                "periods": [
                    {
                        "period": item["period"],
                        "operatingUnitCount": item["operatingUnitCount"],
                        "globalProductCount": item["globalProductCount"],
                        "storeCount": item["storeCount"],
                        "signalFixtureCount": item["signalFixtureCount"],
                    }
                    for item in report["periods"]
                ],
                "canonicalHistoryVerified": report["canonicalHistory"]["verified"],
                "canonicalDataVersions": report["canonicalHistory"]["dataVersions"],
                "historyScanMode": report["canonicalHistory"].get("historyScanMode"),
                "wholeSnapshotRetention": report["canonicalHistory"].get("wholeSnapshotRetention"),
            },
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    return 0 if report["verified"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
