import hashlib
import io
import zipfile

import pytest
from openpyxl import load_workbook

from src.repositories import sqlite_repository
from src.services.competition_era_sample_payload_v2 import sample_contract
from src.services.competition_sample_asset_service import (
    COMPETITION_SAMPLE_ASSET_TABLE,
    ensure_competition_sample_assets,
    get_competition_sample_asset,
    list_competition_sample_asset_metadata,
)
from src.services.competition_sample_report_service import (
    SAMPLE_REPORTS,
    SAMPLE_SHEET_ORDER,
    build_competition_sample_xlsx,
)
from src.services.import_adapter_service import parse_upload_file
from src.services.system_service import RUNTIME_TABLES

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FIXED_CORE_TIME = b"2026-01-01T00:00:00Z"
EXPECTED_SHEET_ROWS = {
    "商品经营明细": 30,
    "店铺经营汇总": 3,
    "流量来源明细": 150,
}


@pytest.fixture()
def isolated_sample_db(monkeypatch, tmp_path):
    monkeypatch.setattr(sqlite_repository, "DB_PATH", tmp_path / "product_workbench.sqlite3")
    monkeypatch.setattr(sqlite_repository, "LOG_DIR", tmp_path)
    monkeypatch.setattr(sqlite_repository, "_WAL_INITIALIZED", False)
    yield tmp_path / "product_workbench.sqlite3"


def test_seed_builder_is_byte_deterministic_and_preserves_era_operating_units():
    for period in (1, 2, 3):
        first = build_competition_sample_xlsx(period)
        second = build_competition_sample_xlsx(period)
        assert first == second
        assert first.startswith(b"PK")

        with zipfile.ZipFile(io.BytesIO(first), "r") as archive:
            core = archive.read("docProps/core.xml")
        assert core.count(FIXED_CORE_TIME) == 2

        workbook = load_workbook(io.BytesIO(first), read_only=True, data_only=True)
        assert workbook.sheetnames == list(SAMPLE_SHEET_ORDER)
        assert workbook["商品经营明细"].max_row == 31
        assert workbook["店铺经营汇总"].max_row == 4
        assert workbook["流量来源明细"].max_row == 151

        rows = list(workbook["商品经营明细"].iter_rows(values_only=True))
        headers = list(rows[0])
        store_idx = headers.index("店铺ID")
        product_idx = headers.index("商品ID")
        sku_idx = headers.index("SKU ID")
        operating_units = {(row[store_idx], row[product_idx], row[sku_idx]) for row in rows[1:]}
        global_products = {row[product_idx] for row in rows[1:]}
        stores = {row[store_idx] for row in rows[1:]}
        assert len(operating_units) == 30
        assert len(global_products) == 10
        assert len(stores) == 3

        contract = sample_contract(period)
        assert contract["operatingProductUnitCount"] == 30
        assert contract["globalProductCount"] == 10
        assert contract["storeCount"] == 3
        assert contract["sheetDataRows"] == EXPECTED_SHEET_ROWS
        assert len(contract["signalOperatingUnits"]) == 3


def test_sqlite_is_runtime_authority_and_seed_is_idempotent(isolated_sample_db):
    first = ensure_competition_sample_assets()
    second = ensure_competition_sample_assets()

    assert first["assetCount"] == 3
    assert first["seededCount"] == 3
    assert first["repairedCount"] == 0
    assert second["assetCount"] == 3
    assert second["seededCount"] == 0
    assert second["repairedCount"] == 0
    assert first["contentSha256ByPeriod"] == second["contentSha256ByPeriod"]
    assert first["runtimeDownloadAuthority"] == "sqlite_blob"
    assert first["runtimeWorkbookGenerationAllowed"] is False
    assert first["releasePrepareRepairAuthority"] is True

    metadata = list_competition_sample_asset_metadata()
    assert [item["period"] for item in metadata] == [1, 2, 3]
    assert all("content" not in item for item in metadata)

    for period in (1, 2, 3):
        asset = get_competition_sample_asset(period)
        payload = asset["content"]
        assert payload.startswith(b"PK")
        assert asset["byteSize"] == len(payload)
        assert asset["contentSha256"] == hashlib.sha256(payload).hexdigest()
        parsed = parse_upload_file(asset["filename"], payload, XLSX_MEDIA_TYPE)
        assert parsed["format"] == "xlsx"
        assert parsed["sheetCount"] == 3
        assert parsed["totalRows"] == 183
        assert {name: len(parsed["sheetRows"][name]) for name in SAMPLE_SHEET_ORDER} == EXPECTED_SHEET_ROWS


def test_runtime_read_fails_closed_but_release_prepare_repairs_stale_seed(isolated_sample_db):
    ensure_competition_sample_assets()
    with sqlite_repository.connect() as conn:
        conn.execute(
            f"""
            UPDATE {COMPETITION_SAMPLE_ASSET_TABLE}
            SET content_blob=?, content_sha256=?, byte_size=?
            WHERE period=1 AND active=1
            """,
            (b"PK-stale-seed-from-interrupted-release", "stale", 999),
        )
        conn.commit()

    with pytest.raises(RuntimeError, match="competition_sample_asset_hash_mismatch"):
        get_competition_sample_asset(1)

    repaired = ensure_competition_sample_assets()
    assert repaired["repairedCount"] == 1
    assert repaired["verifiedCount"] == 3

    asset = get_competition_sample_asset(1)
    payload = asset["content"]
    assert payload.startswith(b"PK")
    assert asset["contentSha256"] == hashlib.sha256(payload).hexdigest()
    assert asset["byteSize"] == len(payload)

    second = ensure_competition_sample_assets()
    assert second["repairedCount"] == 0
    assert second["contentSha256ByPeriod"] == repaired["contentSha256ByPeriod"]


def test_three_signal_fixture_is_downstream_only_and_preserves_expected_trends():
    assert all(len(SAMPLE_REPORTS[p]) == 3 for p in (1, 2, 3))
    conversion_roi = [SAMPLE_REPORTS[p][0]["roi"] for p in (1, 2, 3)]
    scale_roi = [SAMPLE_REPORTS[p][2]["roi"] for p in (1, 2, 3)]
    assert conversion_roi == [3.2, 2.55, 2.08]
    assert scale_roi == [3.85, 4.25, 4.62]
    assert sample_contract(1)["operatingProductUnitCount"] == 30


def test_invalid_period_and_corrupt_upload_fail_closed():
    with pytest.raises(ValueError, match="competition_sample_period_not_found"):
        build_competition_sample_xlsx(4)
    with pytest.raises(ValueError, match="XLSX 文件损坏"):
        parse_upload_file("broken.xlsx", b"not-a-zip", XLSX_MEDIA_TYPE)


def test_sample_assets_are_not_deleted_by_demo_runtime_reset_contract():
    assert COMPETITION_SAMPLE_ASSET_TABLE not in RUNTIME_TABLES


def test_data_import_router_keeps_stable_sample_download_endpoint():
    from src.api.routes import data_import

    paths = {route.path for route in data_import.router.routes}
    assert "/api/data/sample-reports/{period}.xlsx" in paths
